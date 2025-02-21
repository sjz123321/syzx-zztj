import sys
import os
import pandas as pd
import sqlite3
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, 
                           QVBoxLayout, QHBoxLayout, QWidget, QLabel, 
                           QTextEdit, QSplitter, QDialog, QLineEdit,
                           QFormLayout, QDialogButtonBox, QInputDialog)
from PyQt5.QtCore import Qt
from datetime import datetime
from PyQt5.QtWidgets import QComboBox, QMessageBox

class ColumnSelectDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('选择列')
        layout = QFormLayout(self)
        
        self.column_combo = QComboBox(self)
        self.column_combo.addItems(['周次', '日期', '班级', '楼层', '寝室', '类别'])
        
        layout.addRow('选择列:', self.column_combo)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def get_selected_column(self):
        return self.column_combo.currentText()

class AddResponsibleDialog(QDialog):
    def __init__(self, record_id, parent=None):
        super().__init__(parent)
        self.record_id = record_id
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('添加责任人')
        layout = QFormLayout(self)
        
        # 创建输入框
        self.name_input = QLineEdit(self)
        self.points_input = QLineEdit(self)
        
        # 添加输入框到布局
        layout.addRow('责任人姓名:', self.name_input)
        layout.addRow('扣分分值:', self.points_input)
        
        # 添加确定和取消按钮
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

class ExcelProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()  # 先初始化UI
        self.setupDatabase()  # 再设置数据库
        
    def initUI(self):
        self.setWindowTitle('Excel数据处理程序')
        self.setGeometry(100, 100, 1200, 800)
        
        # 创建中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 创建主布局
        main_layout = QHBoxLayout(central_widget)
        
        # 创建左侧控制面板
        left_panel = QWidget()
        self.left_layout = QVBoxLayout(left_panel)
        
        # 创建按钮
        self.process_btn = QPushButton('开始处理Excel文件', self)
        self.process_btn.clicked.connect(self.process_excel_files)  # 修改这里的方法名
        self.process_btn.setMinimumHeight(40)
        
        self.show_data_btn = QPushButton('显示所有数据', self)
        self.show_data_btn.clicked.connect(self.show_all_data)  # 修改这里的方法名
        self.show_data_btn.setMinimumHeight(40)
        
        self.delete_responsible_btn = QPushButton('删除责任人', self)
        self.delete_responsible_btn.clicked.connect(self.delete_responsible)
        self.delete_responsible_btn.setMinimumHeight(40)
        self.left_layout.addWidget(self.delete_responsible_btn)
        
        self.add_responsible_btn = QPushButton('添加责任人', self)
        self.add_responsible_btn.clicked.connect(self.add_responsible)  # 修改这里的方法名
        self.add_responsible_btn.setMinimumHeight(40)
        
        self.filter_btn = QPushButton('筛选', self)
        self.filter_btn.clicked.connect(self.filter_data)
        self.filter_btn.setMinimumHeight(40)
        
        self.filter_unassigned_btn = QPushButton('筛选未扣分', self)
        self.filter_unassigned_btn.clicked.connect(self.filter_unassigned)
        self.filter_unassigned_btn.setMinimumHeight(40)
        
        self.person_stats_btn = QPushButton('统计个人扣分', self)
        self.person_stats_btn.clicked.connect(self.calculate_person_stats)
        self.person_stats_btn.setMinimumHeight(40)
        
        # 添加导出Excel按钮
        self.export_btn = QPushButton('导出到Excel', self)
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setMinimumHeight(40)
        self.left_layout.addWidget(self.export_btn)
        
        # 创建日志区域
        self.log_label = QLabel('处理日志:', self)
        self.log_text = QTextEdit(self)
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(200)
        
        # 将控件添加到左侧布局
        self.left_layout.addWidget(self.process_btn)
        self.left_layout.addWidget(self.show_data_btn)
        self.left_layout.addWidget(self.add_responsible_btn)
        self.left_layout.addWidget(self.log_label)
        self.left_layout.addWidget(self.log_text)
        self.left_layout.addStretch()
        self.left_layout.addWidget(self.filter_btn)
        self.left_layout.addWidget(self.filter_unassigned_btn)
        self.left_layout.addWidget(self.person_stats_btn)
        
        
        # 创建右侧数据显示区域
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        self.data_label = QLabel('数据显示:', self)
        self.data_text = QTextEdit(self)
        self.data_text.setReadOnly(True)
        
        # 将控件添加到右侧布局
        right_layout.addWidget(self.data_label)
        right_layout.addWidget(self.data_text)
        
        # 创建分割器并添加左右面板
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([400, 800])
        
        # 将分割器添加到主布局
        main_layout.addWidget(splitter)
    
    def setupDatabase(self):
        try:
            # 连接SQLite数据库
            self.conn = sqlite3.connect('dormitory_records.db')
            self.cursor = self.conn.cursor()
            
            # 创建主表 - DormitoryRecords
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS DormitoryRecords (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    WeekNum TEXT NOT NULL,
                    RecordDate TEXT NOT NULL,
                    Class TEXT NOT NULL,
                    Floor TEXT NOT NULL,
                    Room TEXT NOT NULL,
                    Category TEXT NOT NULL,
                    DeductPoints REAL NOT NULL,
                    Description TEXT
                )
            ''')
            
            # 创建责任人表 - ResponsiblePersons
            self.cursor.execute('''
                CREATE TABLE IF NOT EXISTS ResponsiblePersons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    record_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    deduct_points REAL NOT NULL,
                    FOREIGN KEY (record_id) REFERENCES DormitoryRecords(id)
                )
            ''')
            
            self.conn.commit()
            self.log_message("数据库连接成功")
        except Exception as e:
            self.log_message(f"数据库连接错误: {str(e)}")

    def log_message(self, message):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.append(f"[{current_time}] {message}")

    def show_all_data(self):  # 改名为 show_all_data
        try:
            self.data_text.clear()
            
            # 使用JOIN查询获取记录和对应的责任人信息
            query = '''
                SELECT 
                    d.id,
                    d.WeekNum, 
                    d.RecordDate, 
                    d.Class, 
                    d.Floor, 
                    d.Room, 
                    d.Category, 
                    d.DeductPoints, 
                    d.Description,
                    GROUP_CONCAT(r.name || '(' || r.deduct_points || '分)') as responsible_info
                FROM DormitoryRecords d
                LEFT JOIN ResponsiblePersons r ON d.id = r.record_id
                GROUP BY d.id
                ORDER BY d.RecordDate DESC, d.Room ASC
            '''
            
            self.cursor.execute(query)
            records = self.cursor.fetchall()
            
            # 准备表头
            header = "ID\t周次\t日期\t班级\t楼层\t寝室\t类别\t扣分\t说明\t责任人(扣分)\n"
            header += "-" * 120 + "\n"
            self.data_text.insertPlainText(header)
            
            # 显示数据
            for record in records:
                responsible_info = record[-1] if record[-1] else "无"
                row = "\t".join(str(item) for item in record[:-1])  # 除去最后一个字段
                row += f"\t{responsible_info}\n"  # 添加责任人信息
                self.data_text.insertPlainText(row)
            
            self.log_message(f"已显示 {len(records)} 条记录")
            
        except Exception as e:
            self.log_message(f"显示数据时出错: {str(e)}")

    def filter_data(self):
        try:
            # 选择列
            column_dialog = ColumnSelectDialog(self)
            if column_dialog.exec_():
                selected_column = column_dialog.get_selected_column()
                
                # 输入筛选内容
                content, ok = QInputDialog.getText(self, '输入筛选内容', 
                    f'请输入要筛选的{selected_column}内容:')
                if ok:
                    # 根据选择的列名获取数据库中的列名
                    column_map = {
                        '周次': 'WeekNum',
                        '日期': 'RecordDate',
                        '班级': 'Class',
                        '楼层': 'Floor',
                        '寝室': 'Room',
                        '类别': 'Category'
                    }
                    
                    db_column = column_map[selected_column]
                    
                    query = f'''
                        SELECT 
                            d.id,
                            d.WeekNum, 
                            d.RecordDate, 
                            d.Class, 
                            d.Floor, 
                            d.Room, 
                            d.Category, 
                            d.DeductPoints, 
                            d.Description,
                            GROUP_CONCAT(r.name || '(' || r.deduct_points || '分)') as responsible_info
                        FROM DormitoryRecords d
                        LEFT JOIN ResponsiblePersons r ON d.id = r.record_id
                        WHERE d.{db_column} = ?
                        GROUP BY d.id
                        ORDER BY d.RecordDate DESC, d.Room ASC
                    '''
                    
                    self.cursor.execute(query, (content,))
                    self.display_query_results("筛选结果")
                    
        except Exception as e:
            self.log_message(f"筛选数据时出错: {str(e)}")
    
    def filter_unassigned(self):
        try:
            query = '''
                SELECT 
                    d.id,
                    d.WeekNum, 
                    d.RecordDate, 
                    d.Class, 
                    d.Floor, 
                    d.Room, 
                    d.Category, 
                    d.DeductPoints, 
                    d.Description,
                    GROUP_CONCAT(r.name || '(' || r.deduct_points || '分)') as responsible_info
                FROM DormitoryRecords d
                LEFT JOIN ResponsiblePersons r ON d.id = r.record_id
                GROUP BY d.id
                HAVING responsible_info IS NULL
                ORDER BY d.RecordDate DESC, d.Room ASC
            '''
            
            self.cursor.execute(query)
            self.display_query_results("未定义责任人的记录")
            
        except Exception as e:
            self.log_message(f"筛选未扣分记录时出错: {str(e)}")
    
    def calculate_person_stats(self):
        try:
            # 获取责任人姓名
            name, ok = QInputDialog.getText(self, '输入责任人姓名', 
                '请输入要统计的责任人姓名:')
            if not ok or not name:
                return

            # 获取起始周次
            start_week, ok = QInputDialog.getText(self, '输入起始周次', 
                '请输入起始周次:')
            if not ok or not start_week:
                return

            # 获取结束周次
            end_week, ok = QInputDialog.getText(self, '输入结束周次', 
                '请输入结束周次:')
            if not ok or not end_week:
                return

            query = '''
                SELECT 
                    d.id,
                    d.WeekNum, 
                    d.RecordDate, 
                    d.Class, 
                    d.Floor, 
                    d.Room, 
                    d.Category, 
                    d.DeductPoints, 
                    d.Description,
                    GROUP_CONCAT(r.name || '(' || r.deduct_points || '分)') as responsible_info,
                    SUM(CASE WHEN r.name = ? THEN r.deduct_points ELSE 0 END) as personal_deduction
                FROM DormitoryRecords d
                LEFT JOIN ResponsiblePersons r ON d.id = r.record_id
                WHERE EXISTS (
                    SELECT 1 FROM ResponsiblePersons r2 
                    WHERE r2.record_id = d.id AND r2.name = ?
                )
                AND d.WeekNum >= ? AND d.WeekNum <= ?
                GROUP BY d.id
                ORDER BY d.WeekNum ASC, d.Room ASC
            '''
            
            self.cursor.execute(query, (name, name, start_week, end_week))
            records = self.cursor.fetchall()
            
            if not records:
                QMessageBox.information(self, '统计结果', 
                    f'在{start_week}周至{end_week}周期间未找到责任人 {name} 的相关记录')
                return
                
            total_deduction = sum(record[-1] for record in records)
            
            # 显示结果
            self.data_text.clear()
            
            # 显示标题和统计信息
            title = f"责任人 {name} 的扣分记录\n"
            title += f"统计周期：第{start_week}周 至 第{end_week}周\n"
            title += f"总扣分：{total_deduction}分\n"
            title += "-" * 120 + "\n"
            self.data_text.insertPlainText(title)
            
            # 显示表头
            header = "ID\t周次\t日期\t班级\t楼层\t寝室\t类别\t扣分\t说明\t责任人(扣分)\n"
            header += "-" * 120 + "\n"
            self.data_text.insertPlainText(header)
            
            # 显示数据
            for record in records:
                responsible_info = record[-2] if record[-2] else "无"  # -2 是责任人信息
                display_record = list(record[:-2])  # 除去最后两个字段（责任人信息和个人扣分）
                row = "\t".join(str(item) for item in display_record)
                row += f"\t{responsible_info}\n"
                self.data_text.insertPlainText(row)
            
            self.log_message(f"已显示 {name} 在第{start_week}周至第{end_week}周的 {len(records)} 条记录，总扣分：{total_deduction}分")
            
        except Exception as e:
            self.log_message(f"统计个人扣分时出错: {str(e)}")

    def closeEvent(self, event):
        try:
            self.conn.close()
            self.log_message("数据库连接已关闭")
        except:
            pass
        event.accept()
    
    def calculate_summary(self):
        try:
            # 选择列
            column_dialog = ColumnSelectDialog(self)
            if column_dialog.exec_():
                selected_column = column_dialog.get_selected_column()
                
                # 输入统计内容
                content, ok = QInputDialog.getText(self, '输入统计内容', 
                    f'请输入要统计的{selected_column}内容:')
                if ok:
                    column_map = {
                        '周次': 'WeekNum',
                        '日期': 'RecordDate',
                        '班级': 'Class',
                        '楼层': 'Floor',
                        '寝室': 'Room',
                        '类别': 'Category'
                    }
                    
                    db_column = column_map[selected_column]
                    
                    query = f'''
                        SELECT 
                            d.id,
                            d.WeekNum, 
                            d.RecordDate, 
                            d.Class, 
                            d.Floor, 
                            d.Room, 
                            d.Category, 
                            d.DeductPoints, 
                            d.Description,
                            GROUP_CONCAT(r.name || '(' || r.deduct_points || '分)') as responsible_info,
                            SUM(d.DeductPoints) as total_deduction,
                            COUNT(*) as record_count
                        FROM DormitoryRecords d
                        LEFT JOIN ResponsiblePersons r ON d.id = r.record_id
                        WHERE d.{db_column} = ?
                        GROUP BY d.id
                        ORDER BY d.RecordDate DESC, d.Room ASC
                    '''
                    
                    self.cursor.execute(query, (content,))
                    records = self.cursor.fetchall()
                    
                    if not records:
                        QMessageBox.information(self, '统计结果', 
                            f'未找到 {selected_column} 为 {content} 的记录')
                        return
                    
                    total_deduction = sum(record[-2] for record in records)
                    record_count = len(records)
                    
                    self.display_query_results(
                        f"{selected_column} 为 {content} 的统计结果\n"
                        f"总记录数：{record_count}，总扣分：{total_deduction}分"
                    )
                    
        except Exception as e:
            self.log_message(f"统计汇总时出错: {str(e)}")
    
    def display_query_results(self, title):
        records = self.cursor.fetchall()
        self.data_text.clear()
        
        # 显示标题
        self.data_text.insertPlainText(f"{title}\n")
        self.data_text.insertPlainText("-" * 120 + "\n")
        
        # 显示表头
        header = "ID\t周次\t日期\t班级\t楼层\t寝室\t类别\t扣分\t说明\t责任人(扣分)\n"
        header += "-" * 120 + "\n"
        self.data_text.insertPlainText(header)
        
        # 显示数据
        for record in records:
            responsible_info = record[-1] if record[-1] else "无"
            display_record = list(record[:-1])  # 除去最后一个字段（责任人信息）
            row = "\t".join(str(item) for item in display_record[:-1])  # 除去最后一个字段（用于合计）
            row += f"\t{responsible_info}\n"
            self.data_text.insertPlainText(row)
        
        self.log_message(f"已显示 {len(records)} 条记录")

    def add_responsible(self):  # 改名为 add_responsible
        try:
            text, ok = QInputDialog.getText(self, '输入记录ID', 
                '请输入要添加责任人的记录ID:')
            if ok:
                record_id = int(text)
                
                # 检查记录是否存在
                self.cursor.execute('''
                    SELECT COUNT(*) FROM DormitoryRecords WHERE id = ?
                ''', (record_id,))
                
                if self.cursor.fetchone()[0] == 0:
                    self.log_message(f"记录ID {record_id} 不存在")
                    return
                
                # 打开添加责任人对话框
                dialog = AddResponsibleDialog(record_id, self)
                if dialog.exec_():
                    name = dialog.name_input.text()
                    points = float(dialog.points_input.text())
                    
                    # 添加责任人记录
                    self.cursor.execute('''
                        INSERT INTO ResponsiblePersons 
                        (record_id, name, deduct_points)
                        VALUES (?, ?, ?)
                    ''', (record_id, name, points))
                    
                    self.conn.commit()
                    self.log_message(f"已为记录 {record_id} 添加责任人: {name}")
                    
                    # 刷新显示
                    self.show_all_data()
                    
        except Exception as e:
            self.log_message(f"添加责任人时出错: {str(e)}")
            
            

    def process_excel_files(self):  # 改名为 process_excel_files
        try:
            excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
            self.log_message(f"找到 {len(excel_files)} 个Excel文件")
            
            for excel_file in excel_files:
                self.log_message(f"正在处理文件: {excel_file}")
                try:
                    df = pd.read_excel(excel_file)
                    data_start_row = None
                    
                    for index, row in df.iterrows():
                        if str(row[0]).strip() == "周次":
                            data_start_row = index + 1
                            break
                    
                    if data_start_row is None:
                        self.log_message(f"文件 {excel_file} 中未找到'周次'标记行")
                        continue
                    
                    for index, row in df.iloc[data_start_row:].iterrows():
                        try:
                            # 检查记录是否已存在
                            check_query = '''
                                SELECT COUNT(*) FROM DormitoryRecords 
                                WHERE WeekNum=? AND RecordDate=? AND Class=? 
                                AND Floor=? AND Room=? AND Category=? 
                                AND DeductPoints=? AND Description=?
                            '''
                            
                            self.cursor.execute(check_query, (
                                str(row[0]),  # 周次
                                str(row[1]),  # 日期
                                str(row[2]),  # 班级
                                str(row[3]),  # 楼层
                                str(row[4]),  # 寝室
                                str(row[5]),  # 类别
                                float(row[6]), # 扣分
                                str(row[7])    # 说明
                            ))
                            
                            if self.cursor.fetchone()[0] == 0:
                                # 插入新记录
                                insert_query = '''
                                    INSERT INTO DormitoryRecords 
                                    (WeekNum, RecordDate, Class, Floor, Room, 
                                     Category, DeductPoints, Description)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                '''
                                
                                self.cursor.execute(insert_query, (
                                    str(row[0]),  # 周次
                                    str(row[1]),  # 日期
                                    str(row[2]),  # 班级
                                    str(row[3]),  # 楼层
                                    str(row[4]),  # 寝室
                                    str(row[5]),  # 类别
                                    float(row[6]), # 扣分
                                    str(row[7])    # 说明
                                ))
                                self.conn.commit()
                                self.log_message(f"已添加新记录: 周次={row[0]}, 寝室={row[4]}")
                            
                        except Exception as e:
                            self.log_message(f"处理行数据时出错: {str(e)}")
                            continue
                            
                except Exception as e:
                    self.log_message(f"处理文件 {excel_file} 时出错: {str(e)}")
                    continue
                    
            self.log_message("所有文件处理完成")
            self.show_all_data()
            
        except Exception as e:
            self.log_message(f"程序执行错误: {str(e)}")
            
    def delete_responsible(self):
        try:
            # 获取记录 ID
            record_id, ok = QInputDialog.getText(self, '输入记录ID', 
                '请输入要删除责任人的记录ID:')
            if not ok:
                return
                
            record_id = int(record_id)
            
            # 检查记录是否存在
            self.cursor.execute('''
                SELECT COUNT(*) FROM DormitoryRecords WHERE id = ?
            ''', (record_id,))
            
            if self.cursor.fetchone()[0] == 0:
                self.log_message(f"记录ID {record_id} 不存在")
                return
            
            # 获取该记录的所有责任人
            self.cursor.execute('''
                SELECT id, name, deduct_points 
                FROM ResponsiblePersons 
                WHERE record_id = ?
            ''', (record_id,))
            
            responsible_persons = self.cursor.fetchall()
            
            if not responsible_persons:
                self.log_message(f"记录ID {record_id} 没有关联的责任人")
                return
                
            # 显示责任人列表供用户选择
            responsible_list = [f"{person[0]}: {person[1]} ({person[2]}分)" for person in responsible_persons]
            responsible_id, ok = QInputDialog.getItem(self, '选择要删除的责任人', 
                '请选择要删除的责任人:', responsible_list, 0, False)
                
            if ok and responsible_id:
                # 从字符串中提取责任人ID
                responsible_id = int(responsible_id.split(':')[0])
                
                # 删除选中的责任人
                self.cursor.execute('''
                    DELETE FROM ResponsiblePersons 
                    WHERE id = ?
                ''', (responsible_id,))
                
                self.conn.commit()
                self.log_message(f"已删除责任人ID: {responsible_id}")
                
                # 刷新显示
                self.show_all_data()
                
        except Exception as e:
            self.log_message(f"删除责任人时出错: {str(e)}")  

    def export_to_excel(self):
        try:
            # 获取当前时间作为文件名的一部分
            current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'export-{current_time}.xlsx'
            
            # 获取当前显示的数据
            data = self.data_text.toPlainText().strip().split('\n')
            
            # 查找真正的表头行的索引
            header_index = -1
            for i, line in enumerate(data):
                if '周次' in line and '日期' in line and '班级' in line:
                    header_index = i
                    break
            
            if header_index == -1:
                raise ValueError("未找到有效的表头行")
    
            # 获取表头
            headers = data[header_index].split('\t')
            
            # 收集统计信息（表头之前的信息）
            summary_info = '\n'.join(data[:header_index])
            
            # 解析数据行
            rows = []
            for line in data[header_index + 1:]:
                if line.startswith('--'):  # 跳过分隔线
                    continue
                # 只处理包含实际数据的行
                if line and not line.startswith('--') and '\t' in line:
                    row_data = line.split('\t')
                    # 确保数据行的列数与表头一致
                    if len(row_data) > len(headers):
                        row_data = row_data[:len(headers)]
                    elif len(row_data) < len(headers):
                        row_data.extend([''] * (len(headers) - len(row_data)))
                    rows.append(row_data)
    
            # 创建Excel文件
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = '数据导出'
            
            # 写入统计信息（表头前的汇总信息）
            current_row = 1
            summary_lines = summary_info.split('\n')
            for line in summary_lines:
                if line.strip():  # 如果不是空行
                    ws.cell(row=current_row, column=1, value=line)
                    # 合并单元格
                    ws.merge_cells(f'A{current_row}:{chr(65+len(headers)-1)}{current_row}')
                    # 设置左对齐
                    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                current_row += 1
            
            current_row += 1  # 空出一行
            
            # 写入表头
            for col, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=header)
            
            # 写入数据
            for row_data in rows:
                current_row += 1
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=current_row, column=col, value=value)
            
            # 设置边框和对齐方式
            thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            
            # 为数据部分添加边框和居中对齐
            data_start_row = len(summary_lines) + 2  # 表头行
            for row in ws.iter_rows(min_row=data_start_row, 
                                max_row=ws.max_row,
                                min_col=1, 
                                max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 调整列宽 - 修复后的版本
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = chr(64 + col)  # A, B, C, ...
                
                # 获取该列所有单元格
                for row in range(data_start_row, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # 考虑表头长度
                header_cell = ws.cell(row=data_start_row, column=col)
                if header_cell.value:
                    max_length = max(max_length, len(str(header_cell.value)))
                
                # 设置列宽
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(filename)
            
            self.log_message(f"数据已导出到文件: {filename}")
            
            # 显示成功消息
            QMessageBox.information(self, '导出成功', 
                f'数据已成功导出到文件:\n{filename}')
                
        except Exception as e:
            self.log_message(f"导出到Excel时出错: {str(e)}")
            QMessageBox.critical(self, '导出错误', 
                f'导出数据时发生错误:\n{str(e)}')

    def closeEvent(self, event):
        try:
            self.conn.close()
            self.log_message("数据库连接已关闭")
        except:
            pass
        event.accept()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelProcessor()
    ex.show()
    sys.exit(app.exec_())